import os
import streamlit as st
import pandas as pd
import sqlite3
import openpyxl
import statsmodels.api as sm
from drive_utils import download_db_from_drive, upload_db_to_drive

# Configuration
DB_FILE_NAME = "ohlc_bbdata.db"
DRIVE_FOLDER_ID = "1ajjaIMmHobK-kU0NxUfk_cBrhy7ZPGmR"
DRIVE_FILE_ID = "1FWoXxyUSgnOZkC7Gxt_Vjpco0G2L1VJo"

# Download DB if missing
if not os.path.exists(DB_FILE_NAME):
    if DRIVE_FILE_ID:
        with st.spinner("Downloading DB from Google Drive..."):
            download_db_from_drive(DRIVE_FILE_ID, DB_FILE_NAME)
            st.success("Database downloaded.")
    else:
        st.warning("No local DB found and no DRIVE_FILE_ID specified.")

#st.title("📈 Stock OHLC Database Update")
# Set Streamlit app to wide layout
st.set_page_config(layout="wide", page_title="📈 Stock OHLC Database App")


# Template download
template_path = "BTH_Template.xlsx"
if os.path.exists(template_path):
    with open(template_path, "rb") as f:
        st.sidebar.download_button("📥 Download BTH Template Sample", f, "BTH_Template_Sample.xlsx")
else:
    st.sidebar.info("ℹ️ BTH Template not found.")

# Mode
mode = st.sidebar.selectbox("Select Mode", ["Update / Create Stock Database", "Read an Existing Database"])

# Functions
def parse_excel(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    datasets = []
    col_index = 1
    while True:
        stock_cell = ws.cell(row=4, column=col_index)
        if stock_cell.value is None:
            break
        stock = stock_cell.value.split()[0]
        data = []
        row_index = 6
        while True:
            row = [ws.cell(row=row_index, column=col_index + i).value for i in range(7)]
            if all(cell is None for cell in row):
                break
            data.append(row)
            row_index += 1
        if data:
            df = pd.DataFrame(data, columns=["Date", "Open", "High", "Low", "Close", "Volume", "Value"])
            df.insert(0, "Stock", stock)
            df["Date"] = pd.to_datetime(df["Date"]).dt.date
            datasets.append(df)
        col_index += 8
    return pd.concat(datasets, ignore_index=True) if datasets else pd.DataFrame()

def save_to_db(df, db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS stock_data (
            Stock TEXT,
            Date TEXT,
            Open REAL,
            High REAL,
            Low REAL,
            Close REAL,
            Volume INTEGER,
            Value REAL,
            VWAP REAL,
            PRIMARY KEY (Stock, Date)
        )
    """)
    existing_keys = pd.read_sql("SELECT Stock, Date FROM stock_data", conn)
    existing_keys["Date"] = pd.to_datetime(existing_keys["Date"]).dt.date
    df_merged = df.merge(existing_keys, on=["Stock", "Date"], how="left", indicator=True)
    new_data = df_merged[df_merged["_merge"] == "left_only"].drop(columns=["_merge"])
    if not new_data.empty:
        new_data["VWAP"] = (new_data["Value"] / new_data["Volume"]).round(4)
        new_data.to_sql("stock_data", conn, if_exists="append", index=False)
    conn.close()
    return new_data

def read_database(db_path):
    if not os.path.exists(db_path):
        st.error(f"Database file `{db_path}` not found.")
        return pd.DataFrame()
    conn = sqlite3.connect(db_path)
    df = pd.read_sql("SELECT * FROM stock_data", conn, parse_dates=["Date"])
    df["Date"] = pd.to_datetime(df["Date"]).dt.date
    conn.close()
    return df

# === Mode: Upload / Create ===
if mode == "Update / Create Stock Database":
    uploaded_file = st.sidebar.file_uploader("Upload Excel File", type=["xlsx"])
    if uploaded_file:
        parsed_df = parse_excel(uploaded_file)
        if not parsed_df.empty:
            parsed_df["VWAP"] = (parsed_df["Value"] / parsed_df["Volume"]).round(4)
            st.subheader("📋 Preview of Uploaded Data with VWAP")
            st.dataframe(parsed_df.head(10))
            if st.sidebar.button("Save to Database"):
                new_rows = save_to_db(parsed_df, DB_FILE_NAME)
                if not new_rows.empty:
                    summary = new_rows.groupby("Stock")["Date"].agg(["min", "max"]).reset_index()
                    summary.columns = ["Stock", "Date From", "Date To"]
                    st.success(f"✅ {len(new_rows)} new records saved.")
                    st.dataframe(summary)
                else:
                    st.info("ℹ️ No new records inserted.")
        else:
            st.warning("Parsed Excel returned no data.")

    if st.sidebar.button("📤 Upload DB to Google Drive"):
        with st.spinner("Uploading to Google Drive..."):
            file_id = upload_db_to_drive(DB_FILE_NAME, DRIVE_FOLDER_ID)
            st.success(f"Uploaded DB to Google Drive. File ID: {file_id}")

# === Mode: Read DB ===
elif mode == "Read an Existing Database":
    if "db_loaded" not in st.session_state:
        st.session_state.db_loaded = False
    if "db_data" not in st.session_state:
        st.session_state.db_data = pd.DataFrame()

    if st.sidebar.button("🔄 Load Database"):
        df = read_database(DB_FILE_NAME)
        if df.empty:
            st.warning("⚠️ No data found.")
            st.session_state.db_loaded = False
        else:
            st.session_state.db_data = df
            st.session_state.db_loaded = True

    if st.session_state.db_loaded:
        df = st.session_state.db_data
        st.sidebar.subheader("📆 Filter Options")
        min_date, max_date = df["Date"].min(), df["Date"].max()
        date_range = st.sidebar.date_input("Select Date Range", [min_date, max_date])

        stocks = df["Stock"].unique().tolist()
        use_input = st.sidebar.checkbox("🔘 Enable Stock List Input", False)

        if use_input:
            raw = st.sidebar.text_input("Enter Stock List (comma-separated)", "")
            if raw.strip():
                input_stocks = [s.strip().upper() for s in raw.split(",")]
                filtered = [s for s in stocks if s.upper() in input_stocks]
                selected_stocks = st.sidebar.multiselect("Select Stocks", filtered, default=filtered)
            else:
                st.sidebar.info("ℹ️ Enter comma-separated stock symbols.")
                selected_stocks = []
        else:
            selected_stocks = st.sidebar.multiselect("Select Stocks", stocks, default=[])

        data_columns = ["Open", "High", "Low", "Close", "Volume", "Value", "VWAP"]
        selected_columns = st.sidebar.multiselect("Select Data Columns", data_columns, default=["Close"])
        valid_columns = [col for col in selected_columns if col in df.columns]

        filtered_df = df[
            (df["Date"] >= date_range[0]) & (df["Date"] <= date_range[1]) & (df["Stock"].isin(selected_stocks))
        ][["Date", "Stock"] + valid_columns].copy()

        if filtered_df.empty:
            st.warning("⚠️ No data to display.")
            st.stop()

        filtered_df.sort_values(["Stock", "Date"], inplace=True)
        filtered_df[valid_columns] = filtered_df.groupby("Stock")[valid_columns].transform(lambda x: x.ffill())
        filtered_df["Date"] = pd.to_datetime(filtered_df["Date"]).dt.strftime("%Y-%m-%d")

        if len(valid_columns) == 1:
            pivot_df = filtered_df.pivot(index="Date", columns="Stock", values=valid_columns[0]).sort_index()
            st.subheader("📑 Filtered Dataset")
            st.dataframe(pivot_df)

            st.sidebar.subheader("📊 Analysis Options")
            selected_stats = st.sidebar.multiselect("Select Analysis", ["Daily Return", "Volatility", "Correlation", "Regression"])

            if "Daily Return" in selected_stats:
                daily_return = pivot_df.pct_change().dropna()
                st.markdown("**📈 Daily Return (% Change)**")
                st.dataframe(daily_return)

            if "Volatility" in selected_stats:
                volatility = daily_return.std()
                st.markdown("**📉 Volatility (Std Dev of Daily Return)**")
                st.dataframe(volatility.to_frame(name="Volatility"))

            if "Correlation" in selected_stats:
                correlation = daily_return.corr()
                st.markdown("**🔗 Correlation Matrix**")
                st.dataframe(correlation)

            if "Regression" in selected_stats:
                st.markdown("**📐 Regression vs Benchmark**")
                benchmark = st.selectbox("Select Benchmark", pivot_df.columns)
                regressions = []
                for stock in daily_return.columns:
                    if stock == benchmark:
                        continue
                    X = sm.add_constant(daily_return[benchmark])
                    y = daily_return[stock]
                    model = sm.OLS(y, X).fit()
                    regressions.append({
                        "Stock": stock,
                        "Beta": model.params[benchmark],
                        "Alpha": model.params["const"],
                        "R²": model.rsquared
                    })
                st.dataframe(pd.DataFrame(regressions))

        else:
            reshaped = []
            for col in valid_columns:
                temp = filtered_df.pivot(index="Date", columns="Stock", values=col)
                temp.columns = [f"{stock}_{col}" for stock in temp.columns]
                reshaped.append(temp)
            combined = pd.concat(reshaped, axis=1).sort_index()
            st.subheader("📑 Filtered Dataset")
            st.dataframe(combined)
