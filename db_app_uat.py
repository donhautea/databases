# db_app_uat.py (with integrated equity_monitor.py logic)

from drive_utils import download_db_from_drive, upload_db_to_drive
import streamlit as st
import pandas as pd
import openpyxl
import sqlite3
import os

import matplotlib.pyplot as plt
import statsmodels.api as sm
from datetime import date

# --- Streamlit App Config ---
st.set_page_config(page_title="📈 Stock Volume & Equity Monitor", layout="wide")

# === Main Menu ===
st.sidebar.title("📋 Main Menu")
main_mode = st.sidebar.radio("Select App Section", ["📘 Stock DB Manager", "📊 Equity Monitor"])

# === EQUITY MONITOR ===
if main_mode == "📊 Equity Monitor":
    st.title("📊 Equity Database Monitoring")

    st.sidebar.header("Upload Excel File")
    uploaded_file = st.sidebar.file_uploader("Choose an Excel (.xlsx) file", type="xlsx")

    if not uploaded_file:
        st.info("📥 Please upload an Excel file with the required sheet structure to begin.")
        st.stop()

    fund_sheet_map = {
        "SSS": ["SSS_FVTPL", "SSS_FVTOCI"],
        "EC": ["EC_FVTPL", "EC_FVTOCI"],
        "MPF": ["MPF_FVTPL"],
        "NVPF": ["NVPF_FVTPL"]
    }
    all_funds = list(fund_sheet_map.keys())

    sheets = pd.read_excel(uploaded_file, sheet_name=None)
    selected_fund = st.sidebar.radio("Select Fund to Analyze:", all_funds + ["All Funds"])
    date_from = st.sidebar.date_input("Date From", date.today())
    date_to = st.sidebar.date_input("Date To", date.today())
    show_custom_summary = st.sidebar.checkbox("Show Net Value Summary", value=True)
    chart_fund = st.sidebar.checkbox("Bar Chart by Fund: Total Value by Buy/Sell")
    chart_stock = st.sidebar.checkbox("Bar Chart by Fund: Buy/Sell by Stock")
    chart_broker = st.sidebar.checkbox("Bar Chart by Broker: Buy/Sell by Value")

    date_period = f"{date_from} to {date_to}" if date_from != date_to else f"{date_from}"

    dfs = []
    funds = all_funds if selected_fund == "All Funds" else [selected_fund]
    required = {"Date", "Classification", "Stock", "Buy_Sell", "Broker", "Volume", "Price"}

    for fund in funds:
        for sheet in fund_sheet_map[fund]:
            df = sheets.get(sheet)
            if df is None or not required.issubset(df.columns):
                continue
            df = df.copy()
            df["Date"] = pd.to_datetime(df["Date"]).dt.date
            df = df[df["Date"].between(date_from, date_to)]
            df["Value"] = df["Volume"] * df["Price"]
            df["Fund"] = fund
            dfs.append(df)

    if not dfs:
        st.warning("No valid data for the selected fund(s) and date range.")
        st.stop()

    full_df = pd.concat(dfs, ignore_index=True)

    def fmt_value_millions(x):
        return f"₱{x:,.1f}M" if x != 0 else ""

    st.subheader(f"📁 Data for: {selected_fund}")
    full_df_display = full_df.copy()
    full_df_display["Value"] = full_df_display["Value"].apply(lambda x: f"₱{x:,.2f}")
    st.dataframe(full_df_display)

    if show_custom_summary:
        st.subheader("📊 Summary by Fund: Net Value")
        st.markdown(f"**🗓️ Period: {date_period}**")

        grouped = full_df.groupby(["Fund", "Buy_Sell"])["Value"].sum().unstack().fillna(0)
        grouped = grouped.rename(columns={"B": "Buy Value", "S": "Sell Value"})
        grouped["Net Value"] = grouped.get("Buy Value", 0.0) - grouped.get("Sell Value", 0.0)
        grouped["% Distribution"] = ((grouped["Buy Value"] + grouped["Sell Value"]) / 
                                     (grouped["Buy Value"] + grouped["Sell Value"]).sum()) * 100
        grouped = grouped.reset_index()

        summary_df = pd.concat([
            grouped,
            pd.DataFrame([{
                "Fund": "Total",
                "Buy Value": grouped["Buy Value"].sum(),
                "Sell Value": grouped["Sell Value"].sum(),
                "Net Value": grouped["Net Value"].sum(),
                "% Distribution": 100.0
            }])
        ], ignore_index=True)

        def fmt_currency(val): return f"₱{val:,.2f}"
        def fmt_net(val): return f"<span style='color:{'green' if val >= 0 else 'red'}'>₱{val:,.2f}</span>"
        def fmt_pct(val): return f"{val:,.2f}%"

        summary_df["Buy Value"] = summary_df["Buy Value"].apply(fmt_currency)
        summary_df["Sell Value"] = summary_df["Sell Value"].apply(fmt_currency)
        summary_df["Net Value"] = summary_df["Net Value"].apply(fmt_net)
        summary_df["% Distribution"] = summary_df["% Distribution"].apply(fmt_pct)

        st.markdown(summary_df.to_html(escape=False, index=False), unsafe_allow_html=True)

    if chart_fund:
        st.subheader("Bar Chart: Total Value by Fund & Buy/Sell")
        cd = full_df.groupby(["Fund", "Buy_Sell"])["Value"].sum().unstack().fillna(0) / 1e6
        ax = cd.plot(kind="bar", figsize=(10, 5), title=f"Total Value by Fund (₱M) — {date_period}")
        ax.set_ylabel("₱ Millions")
        for cont in ax.containers:
            ax.bar_label(cont, labels=[fmt_value_millions(v) for v in cont.datavalues], fontsize=8)
        st.pyplot(plt)

    if chart_stock:
        st.subheader("Bar Chart: Buy/Sell by Stock")
        stocks = sorted(full_df["Stock"].unique())
        sel = st.multiselect("Select Stocks:", stocks, default=stocks)
        if sel:
            fd = full_df[full_df["Stock"].isin(sel)]
            cd = fd.groupby(["Stock", "Buy_Sell"])["Value"].sum().unstack().fillna(0) / 1e6
            ax = cd.plot(kind="bar", figsize=(12, 6), title=f"Buy/Sell by Stock (₱M) — {date_period}")
            ax.set_ylabel("₱ Millions")
            for cont in ax.containers:
                ax.bar_label(cont, labels=[fmt_value_millions(v) for v in cont.datavalues], fontsize=8)
            st.pyplot(plt)

    if chart_broker:
        st.subheader("Bar Chart: Buy/Sell by Broker")
        cd = full_df.groupby(["Broker", "Buy_Sell"])["Value"].sum().unstack().fillna(0) / 1e6
        ax = cd.plot(kind="bar", figsize=(12, 6), title=f"Buy/Sell by Broker (₱M) — {date_period}")
        ax.set_ylabel("₱ Millions")
        for cont in ax.containers:
            ax.bar_label(cont, labels=[fmt_value_millions(v) for v in cont.datavalues], fontsize=8)
        st.pyplot(plt)

    st.subheader("📘 Weighted Average Price by Fund, Stock, Buy/Sell")
    weighted_summary = full_df.groupby(["Date", "Fund", "Buy_Sell", "Stock"]).agg(
        Total_Volume=("Volume", "sum"),
        Total_Value=("Value", "sum")
    ).reset_index()
    weighted_summary["Weighted_Avg_Price"] = (weighted_summary["Total_Value"] / weighted_summary["Total_Volume"]).round(2)
    weighted_summary["Total_Value"] = weighted_summary["Total_Value"].round(2)
    weighted_summary["Total_Volume"] = weighted_summary["Total_Volume"].apply(lambda x: f"{x:,.0f}")
    weighted_summary["Total_Value"] = weighted_summary["Total_Value"].apply(lambda x: f"₱{x:,.2f}")
    weighted_summary["Weighted_Avg_Price"] = weighted_summary["Weighted_Avg_Price"].apply(lambda x: f"₱{x:,.2f}")
    st.dataframe(weighted_summary, use_container_width=True)

# === STOCK DB MANAGER ===
elif main_mode == "📘 Stock DB Manager":

    # App Config
    st.set_page_config(layout="wide", page_title="📈 Stock OHLC Database Manager")
    
    # Constants
    DB_FILE_NAME = "ohlc_bbdata.db"
    DRIVE_FOLDER_ID = "1ajjaIMmHobK-kU0NxUfk_cBrhy7ZPGmR"
    DRIVE_FILE_ID = "1FWoXxyUSgnOZkC7Gxt_Vjpco0G2L1VJo"
    
    # Download DB from Google Drive if missing
    if not os.path.exists(DB_FILE_NAME) and DRIVE_FILE_ID:
        with st.spinner("🔄 Downloading DB from Google Drive..."):
            download_db_from_drive(DRIVE_FILE_ID, DB_FILE_NAME)
            st.success("✅ Database downloaded.")
    
    # Title
    st.title("📊 Stock OHLC Database Management")
    
    # Template download
    template_path = "data/Integrated_BTH_Template.xlsx"
    if os.path.exists(template_path):
        with open(template_path, "rb") as f:
            st.sidebar.download_button("📥 Download BTH Template Sample", f, "BTH_Template_Sample.xlsx")
    
    # Sidebar Mode
    mode = st.sidebar.radio("Select Mode", ["Update / Create Stock Database", "Read an Existing Database"])
    db_path = DB_FILE_NAME
    
    
    # Functions
    def parse_excel(file):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        datasets = []
        col_index = 1
        while True:
            stock_cell = ws.cell(row=4, column=col_index)
            if stock_cell.value is None:
                break
            stock = stock_cell.value.split()[0]
            data = []
            row_index = 6
            while True:
                row = [
                    ws.cell(row=row_index, column=col_index + i).value for i in range(7)
                ]
                if all(cell is None for cell in row):
                    break
                row = [None if str(cell).strip().upper() in ["#N/A", "N/A", "#N/A N/A"] else cell for cell in row]
                data.append(row)
                row_index += 1
            if data:
                df = pd.DataFrame(data, columns=["Date", "Open", "High", "Low", "Close", "Volume", "Value"])
                df.insert(0, "Stock", stock)
                df["Date"] = pd.to_datetime(df["Date"]).dt.date
                datasets.append(df)
            col_index += 8
        return pd.concat(datasets, ignore_index=True) if datasets else pd.DataFrame()
    
    def save_to_db(df, db_path):
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS stock_data (
                Stock TEXT,
                Date TEXT,
                Open REAL,
                High REAL,
                Low REAL,
                Close REAL,
                Volume INTEGER,
                Value REAL,
                VWAP REAL,
                PRIMARY KEY (Stock, Date)
            )
        """)
    
        
        # Clean and convert Date
        df = df.dropna(subset=["Stock", "Date"]).copy()
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
        df = df.dropna(subset=["Date"])  # Remove rows with invalid Date
        df = df[df["Date"] != pd.to_datetime("1970-01-01").date()]  # Explicitly remove 1970-01-01
    
        if df.empty:
            conn.close()
            return pd.DataFrame()  # Nothing to insert
    
        # Calculate VWAP
        df["VWAP"] = (df["Value"] / df["Volume"]).round(4)
    
        # Read existing keys
        existing_keys = pd.read_sql("SELECT Stock, Date FROM stock_data", conn)
        existing_keys["Date"] = pd.to_datetime(existing_keys["Date"]).dt.date
    
        # Filter out duplicates
        df_merged = df.merge(existing_keys, on=["Stock", "Date"], how="left", indicator=True)
        new_data = df_merged[df_merged["_merge"] == "left_only"].drop(columns=["_merge"])
    
        if not new_data.empty:
            new_data = new_data.drop_duplicates(subset=["Stock", "Date"])
            new_data.to_sql("stock_data", conn, if_exists="append", index=False)
    
        conn.close()
        return new_data
    
    
    
    def read_database(db_path):
        if not os.path.exists(db_path):
            st.error("❌ Database not found.")
            return pd.DataFrame()
        conn = sqlite3.connect(db_path)
        df = pd.read_sql("SELECT * FROM stock_data", conn, parse_dates=["Date"])
        conn.close()
    
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
        df = df.dropna(subset=["Date"])
        df = df[df["Date"] != pd.to_datetime("1970-01-01").date()]  # Filter again here for safety
    
        return df
    
    
    # Main Logic
    if mode == "Update / Create Stock Database":
        uploaded_file = st.sidebar.file_uploader("Upload Excel File", type=["xlsx"])
        if uploaded_file:
            parsed_df = parse_excel(uploaded_file)
            if not parsed_df.empty:
                parsed_df["VWAP"] = (parsed_df["Value"] / parsed_df["Volume"]).round(4)
                st.subheader("📋 Preview of Uploaded Data with VWAP")
                st.dataframe(parsed_df.head(10))
                if st.sidebar.button("💾 Save to Database"):
                    new_rows = save_to_db(parsed_df, db_path)
                    if not new_rows.empty:
                        summary = new_rows.groupby("Stock")["Date"].agg(["min", "max"]).reset_index()
                        summary.columns = ["Stock", "Date From", "Date To"]
                        st.success(f"✅ {len(new_rows)} new records saved.")
                        st.dataframe(summary)
                    else:
                        st.info("ℹ️ No new records inserted.")
    
        if st.sidebar.button("📤 Upload DB to Google Drive"):
            with st.spinner("Uploading to Google Drive..."):
                file_id = upload_db_to_drive(db_path, DRIVE_FOLDER_ID)
                st.success(f"✅ Uploaded DB to Drive. File ID: {file_id}")
    
        # --- USER CREDENTIALS (Hardcoded for now, can be replaced with DB verification) ---
        AUTHORIZED_USERS = {
            "admin": "08201977",
            "geonel": "miguel",
        }
        
        # --- AUTHENTICATION LOGIC ---
        st.sidebar.markdown("---")
        st.sidebar.subheader("🔐 Admin Login to Delete Records")
        
        username = st.sidebar.text_input("Username")
        password = st.sidebar.text_input("Password", type="password")
        
        is_authenticated = username in AUTHORIZED_USERS and password == AUTHORIZED_USERS[username]
        
        if is_authenticated:
            st.sidebar.success(f"Welcome, {username}. You may proceed.")
            
            st.sidebar.markdown("---")
            st.sidebar.subheader("🗑️ Delete from Database")
            delete_type = st.sidebar.selectbox("Delete by:", ["None", "Date", "Stock and Date"])
            if delete_type != "None":
                with sqlite3.connect(db_path) as conn:
                    if delete_type == "Date":
                        date_to_delete = st.sidebar.date_input("Select Date to Delete")
                        if st.sidebar.button("Delete Records by Date"):
                            deleted = conn.execute("DELETE FROM stock_data WHERE Date = ?", (str(date_to_delete),)).rowcount
                            st.sidebar.success(f"✅ {deleted} records deleted for {date_to_delete}")
                    elif delete_type == "Stock and Date":
                        stock_input = st.sidebar.text_input("Stock Symbol (e.g. AC)")
                        date_input = st.sidebar.date_input("Select Date")
                        if st.sidebar.button("Delete Record for Stock and Date"):
                            deleted = conn.execute(
                                "DELETE FROM stock_data WHERE Stock = ? AND Date = ?", (stock_input, str(date_input))
                            ).rowcount
                            st.sidebar.success(f"✅ {deleted} record(s) deleted for {stock_input} on {date_input}")
                            
        else:
            if username or password:
                st.sidebar.error("❌ Invalid credentials.")
            st.sidebar.info("Please log in to access delete functions.")
    
    
    elif mode == "Read an Existing Database":
        df = read_database(db_path)
        if not df.empty:
            st.sidebar.markdown("---")
            st.sidebar.subheader("📆 Filter Options")
            min_date, max_date = df["Date"].min(), df["Date"].max()
            date_range = st.sidebar.date_input("Select Date Range", [min_date, max_date], min_value=min_date, max_value=max_date)
    
            stocks = sorted(df["Stock"].unique())
            stock_input_mode = st.sidebar.checkbox("🔘 Use Stock List Input")
            if stock_input_mode:
                raw_input = st.sidebar.text_input("Enter comma-separated stocks (e.g. AC, ALI)")
                input_stocks = [s.strip().upper() for s in raw_input.split(",") if s.strip()]
                selected_stocks = [s for s in stocks if s in input_stocks]
            else:
                selected_stocks = st.sidebar.multiselect("Select Stocks", stocks, default=stocks[:5])
    
            columns = ["Open", "High", "Low", "Close", "Volume", "Value", "VWAP"]
            selected_columns = st.sidebar.multiselect("Select Columns", columns, default=["Close"])
    
            filtered_df = df[
                (df["Date"] >= date_range[0]) & (df["Date"] <= date_range[1]) & (df["Stock"].isin(selected_stocks))
            ].copy()
    
            if filtered_df.empty:
                st.warning("⚠️ No matching records.")
                st.stop()
    
            filtered_df.sort_values(["Stock", "Date"], inplace=True)
            filtered_df[selected_columns] = filtered_df.groupby("Stock")[selected_columns].transform(lambda x: x.ffill())
            pivot_df = filtered_df.pivot(index="Date", columns="Stock", values=selected_columns[0])
            pivot_df.index = pd.to_datetime(pivot_df.index).strftime("%Y-%m-%d")
    
            st.subheader("📑 Filtered Dataset")
            st.dataframe(pivot_df)
    
            st.sidebar.markdown("---")
            st.sidebar.subheader("📊 Analysis Options")
            selected_analyses = st.sidebar.multiselect("Select Analyses", ["Daily Return", "Volatility", "Correlation", "Regression"])
            if selected_analyses:
                st.subheader("📈 Analysis Results")
    
            if "Daily Return" in selected_analyses:
                daily_return = pivot_df.pct_change().dropna()
                st.markdown("**📈 Daily Return**")
                st.dataframe(daily_return)
    
            if "Volatility" in selected_analyses:
                volatility = daily_return.std()
                st.markdown("**📉 Volatility (Std Dev)**")
                st.dataframe(volatility.to_frame(name="Volatility"))
    
            if "Correlation" in selected_analyses:
                correlation = daily_return.corr()
                st.markdown("**🔗 Correlation Matrix**")
                st.dataframe(correlation)
    
            if "Regression" in selected_analyses:
                benchmark = st.selectbox("Benchmark Stock", pivot_df.columns)
                reg_results = []
                for stock in pivot_df.columns:
                    if stock == benchmark:
                        continue
                    y = daily_return[stock].dropna()
                    x = daily_return[benchmark].reindex(y.index).dropna()
                    x = sm.add_constant(x)
                    model = sm.OLS(y.loc[x.index], x).fit()
                    reg_results.append({
                        "Stock": stock,
                        "Alpha": model.params["const"],
                        "Beta": model.params[benchmark],
                        "R²": model.rsquared
                    })
                st.markdown("**📐 Regression Stats vs. Benchmark**")
                st.dataframe(pd.DataFrame(reg_results))
